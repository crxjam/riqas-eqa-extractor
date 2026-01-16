import re
from datetime import datetime, date

from pathlib import Path
from pdfminer.high_level import extract_text

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from dataclasses import dataclass
from typing import List, Dict, Optional

# ============================
# RCPA PARSER (Automated DIFF)
# ============================

@dataclass
class RCPARow:
    program: str
    participant_id: Optional[str]
    survey_no: Optional[int]
    report_date: Optional[date]
    sample_id: str
    analyte: str
    your_result: float
    expected_result: float
    review: str
    z_score: float
    aps_score: float

def is_rcpa_report(text: str) -> bool:
    t = text.lower()
    return ("rcpa" in t) or ("rcpaqap" in t) or ("issue date:" in t and "participant id:" in t)

def coerce_cycle_history_types(hist_df: pd.DataFrame) -> pd.DataFrame:
    if hist_df is None or hist_df.empty:
        return hist_df

    # dates
    if "Report_Date" in hist_df.columns:
        hist_df["Report_Date"] = pd.to_datetime(hist_df["Report_Date"], errors="coerce")

    # numerics used in comparisons / thresholds
    numeric_cols = [
        "diff_from_mean", "SDI", "%DEV", "Target_Score",
        "TEa_or_TDPA(%)", "Internal_TEa", "Peer_Mean", "Your_Result"
    ]
    for col in numeric_cols:
        if col in hist_df.columns:
            hist_df[col] = pd.to_numeric(hist_df[col], errors="coerce")

    return hist_df



def parse_rcpa_participant_id(text: str) -> Optional[str]:
    m = re.search(r"Participant ID:\s*([A-Z]{1,3}/\d+(?:\.\d+)?)", text)
    return m.group(1).strip() if m else None

def parse_rcpa_program(text: str) -> Optional[str]:
    for line in text.splitlines()[:120]:
        if "Automated Differential" in line:
            return line.strip()
    return None

def parse_rcpa_survey_no(text: str) -> Optional[int]:
    m = re.search(r"\bSurvey\s*:?[\s]*(\d+)\b", text)
    return int(m.group(1)) if m else None

def parse_rcpa_report_issue_date(text: str) -> Optional[date]:
    m = re.search(
        r"Report Issue Date:\s*(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})",
        text,
        re.IGNORECASE
    )
    if not m:
        return None

    s = m.group(1).strip()

    # try abbreviated month first ("08 Jan 2026")
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    return None


def parse_rcpa_issue_date(text: str) -> Optional[date]:
    # Backwards-compatible alias used by older parts of the code
    return parse_rcpa_report_issue_date(text)


def parse_rcpa_summary_of_performance(text: str) -> List[RCPARow]:
    """
    Token-based parser for RCPA 'Summary of Performance' table.
    Works when pdfminer breaks the row into multiple lines / columns.
    Expects 2 samples per analyte (left + right).
    """

    program = parse_rcpa_program(text) or "RCPA Automated Differential"
    participant_id = parse_rcpa_participant_id(text)
    survey_no = parse_rcpa_survey_no(text)
    report_date = parse_rcpa_report_issue_date(text) or parse_rcpa_issue_date(text)

    # -------- find the two sample IDs anywhere in the report ----------
    sample_ids = []
    for m in re.finditer(r"Sample:\s*(HA-[A-Z0-9]+-\d{2}-\d{2})", text, flags=re.IGNORECASE):
        sample_ids.append(m.group(1).strip())

    # de-dupe preserve order
    seen = set()
    sample_ids = [s for s in sample_ids if not (s in seen or seen.add(s))]

    if len(sample_ids) < 2:
        # fallback: any HA-... token
        for m in re.finditer(r"(HA-[A-Z0-9]+-\d{2}-\d{2})", text, flags=re.IGNORECASE):
            sample_ids.append(m.group(1).strip())
        seen = set()
        sample_ids = [s for s in sample_ids if not (s in seen or seen.add(s))]

    if len(sample_ids) < 2:
        return []

    left_sample, right_sample = sample_ids[0], sample_ids[1]

    # -------- locate the summary section (best-effort) ----------
    low = text.lower()
    start = low.find("summary of performance")
    if start == -1:
        # some PDFs use slightly different heading
        m = re.search(r"summary\s+of\s+performance", text, flags=re.IGNORECASE)
        if not m:
            return []
        start = m.start()

    # take a window after the section start (big enough to contain the table)
    window = text[start:start + 12000]

    # normalize lines: keep order, drop empties
    lines = [ln.strip() for ln in window.splitlines() if ln.strip()]

    # We’ll parse per analyte by scanning tokens after the analyte heading.
    # This is robust against column wrapping.


############################################
# 1. PDF PARSING
############################################

from pdfminer.layout import LAParams

def read_pdf_text(pdf_path):
    """Portable text extraction using pdfminer.six (pure Python), tuned for tables."""
    laparams = LAParams(
        all_texts=True,
        detect_vertical=True,
        line_margin=0.15,
        word_margin=0.1,
        char_margin=2.0,
        boxes_flow=None,  # IMPORTANT: stops pdfminer re-ordering columns
    )
    text = extract_text(str(pdf_path), laparams=laparams) or ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text

def normalize_pdf_text(txt: str) -> str:
    # collapse weird whitespace inside lines, keep newlines
    lines = [re.sub(r"[ \t]+", " ", ln).strip() for ln in txt.splitlines()]
    # drop empty noise lines
    lines = [ln for ln in lines if ln]
    return "\n".join(lines)






def parse_metadata(full_text: str) -> dict:
    """
    Pull top-level metadata like Cycle, Sample, Date, Lab name, Instrument, Panel.
    These strings are present in the PDF header/intro.

    Example we look for:
    "CYCLE 18    SAMPLE 10    13/10/2025"
    "MONTHLY HAEMATOLOGY"
    "NHLS Red Cross Childrens Hospital"
    """
    # Cycle / Sample / Date
    m = re.search(r"CYCLE\s+(\d+)\s+SAMPLE\s+(\d+)\s+(\d{2}/\d{2}/\d{4})", full_text)
    if m:
        cycle_no = int(m.group(1))
        sample_no = int(m.group(2))
        report_date_str = m.group(3)
        report_date = datetime.strptime(report_date_str, "%d/%m/%Y").date()
    else:
        cycle_no = None
        sample_no = None
        report_date = None

    # Panel / Scheme name e.g. "MONTHLY HAEMATOLOGY"
    panel_match = re.search(r"(MONTHLY\s+[A-Z ]+)", full_text)
    panel_name = panel_match.group(1).strip() if panel_match else None

    # Lab name: first meaningful line usually near top of page 1
    # We'll take the first line that contains "NHLS" or "Hospital"
    lab_name = None
    for line in full_text.splitlines()[:50]:
        line_clean = line.strip()
        if len(line_clean) < 3:
            continue
        if "NHLS" in line_clean or "Hospital" in line_clean or "HOSPITAL" in line_clean:
            lab_name = line_clean
            break

    # Instrument / Method: often visible near header as something like "Sysmex XN2000"
    # We'll try to grab a Sysmex / Roche / Beckman string if present.
    instrument_match = re.search(
        r"(Sysmex\s+[A-Za-z0-9\-/ ]+|Beckman\s+Coulter\s+[A-Za-z0-9\-/ ]+|Roche\s+[A-Za-z0-9\-/ ]+)",
        full_text[:2000]
    )
    instrument_name = instrument_match.group(1).strip() if instrument_match else None

    return {
        "cycle_no": cycle_no,
        "sample_no": sample_no,
        "report_date": report_date,
        "panel_name": panel_name,
        "lab_name": lab_name,
        "instrument_name": instrument_name,
    }


def extract_tdpa(full_text: str, idx: int) -> Optional[float]:
    """
    TEa / performance tolerance usually appears on the page as:
    'TDPA = 3.7%'
    We'll search ~2000 chars after the analyte label first.
    """
    window = full_text[idx:idx + 2500]
    m = re.search(r"TDPA\s*=\s*([0-9.]+)%", window)
    if m:
        return float(m.group(1))
    # fallback: look backwards
    window_back = full_text[max(0, idx - 2500):idx]
    m2 = re.search(r"TDPA\s*=\s*([0-9.]+)%", window_back)
    if m2:
        return float(m2.group(1))
    return None


from typing import Optional, Dict

def parse_single_analyte(full_text: str, analyte_label: str) -> Optional[Dict]:
    idx = full_text.find(analyte_label)
    if idx == -1:
        return None

    # Work on a window AFTER the analyte label (that’s where the keyed fields are)
    post = full_text[idx: idx + 2000]

    # Your Result and SDI appear on the same line in your sample:
    # "Your Result 353.000 SDI -6.11"
    m_yr_sdi = re.search(r"Your Result\s+([-\d.]+)\s+SDI\s+([-\d.]+)", post)
    if not m_yr_sdi:
        return None
    yr = float(m_yr_sdi.group(1))
    sdi = float(m_yr_sdi.group(2))

    # Mean for Comparison and TS appear on the same line:
    # "Mean for Comparison 521.531 TS 10"
    m_mean_ts = re.search(r"Mean for Comparison\s+([-\d.]+)\s+TS\s+([-\d.]+)", post)
    if not m_mean_ts:
        return None
    mean_comp = float(m_mean_ts.group(1))
    ts = float(m_mean_ts.group(2))

    # %DEV line:
    # "%DEV -32.3 Sample Number"  (so take the number right after %DEV)
    m_dev = re.search(r"%DEV\s+([-\d.]+)", post)
    dev = float(m_dev.group(1)) if m_dev else None

    # TDPA:
    m_tdpa = re.search(r"TDPA\s*=\s*([0-9.]+)%", post)
    tdpa = float(m_tdpa.group(1)) if m_tdpa else None

    return {
        "AnalyteRaw": analyte_label,
        "Peer Mean": mean_comp,
        "Your Result": yr,
        "%DEV": dev,
        "SDI": sdi,
        "Target Score": ts,
        "TDPA_limit_percent": tdpa,
    }

def _norm_analyte_name(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

def find_analyte_labels(full_text: str) -> list[str]:
    # Grab whatever comes after "Mean for Comparison" either on SAME line or NEXT line
    # then clean/filter.
    raw = re.findall(r"Mean for Comparison\s*\n?([^\n]+)", full_text, flags=re.IGNORECASE)

    labels = []
    for cand in raw:
        cand = cand.strip()

        # drop obvious junk
        if not cand or len(cand) < 3:
            continue
        low = cand.lower()
        if low.startswith("laboratory ref") or low.startswith("cycle") or low.startswith("all methods"):
            continue

        # must look like an analyte with a unit-ish pattern
        # e.g. "Calcium, mmol/l" "Amylase, U/l @ 37°C" "Protein, Total, g/l"
        if re.search(r",\s*[^,]{1,25}/[^,]{1,25}", cand) or re.search(r",\s*[^,]+@",
                                                                       cand):
            labels.append(cand)

    # de-dupe preserve order
    seen = set()
    out = []
    for x in labels:
        if x not in seen:
            seen.add(x)
            out.append(x)
    # ------------------------------------------------------------
    # Fallback parser for RCPA "Summary of Performance" tables
    # Works for Auto Diff AND COAG factors by reading each row line,
    # grabbing the analyte name before the first number, and mapping
    # the numeric columns for Sample 1 + Sample 2 (and optional MPS).
    # ------------------------------------------------------------
    if not out:
        # Identify the table header line (best-effort), then parse rows until "Overall Performance"
        start_idx = 0
        for idx, ln in enumerate(lines):
            l = ln.lower()
            if ("test" in l and "your result" in l and "expected result" in l):
                start_idx = idx + 1
                break

        for ln in lines[start_idx:]:
            if "overall performance" in ln.lower():
                break

            # Extract all numeric tokens from the line (handles +/-, decimals)
            nums = re.findall(r"[-+]?\d+(?:\.\d+)?", ln)
            if len(nums) < 4:
                # likely an interpretation row ("Normal", "Abnormal", "No Target Set") → skip
                continue

            # Get analyte/test name: everything before the first numeric occurrence
            m_first = re.search(r"[-+]?\d+(?:\.\d+)?", ln)
            if not m_first:
                continue
            analyte = ln[:m_first.start()].strip()
            if not analyte:
                continue

            def f(i):
                return float(nums[i]) if i < len(nums) else None

            # Default mapping
            # Many reports contain:
            # Sample 1: Your, Expected, Z, APS  (4 numbers)
            # Sample 2: Your, Expected, Z, APS  (4 numbers)
            # Optional: MPS (1 number)
            y1 = f(0)
            e1 = f(1)

            z1 = f(2) if len(nums) >= 3 else None
            a1 = f(3) if len(nums) >= 4 else None

            y2 = f(4) if len(nums) >= 5 else None
            e2 = f(5) if len(nums) >= 6 else None

            z2 = f(6) if len(nums) >= 7 else None
            a2 = f(7) if len(nums) >= 8 else None

            # Optional MPS at the end (COAG factors often has this)
            mps = f(8) if len(nums) >= 9 else None

            # If we only got 4 numbers, assume it's: y1, e1, y2, e2 (no z/aps)
            if len(nums) == 4:
                z1 = a1 = z2 = a2 = None
                y2 = f(2)
                e2 = f(3)

            # Basic review extraction (optional, but nice)
            review = None
            low = ln.lower()
            if "within aps" in low:
                review = "Within APS"
            elif "high" in low:
                review = "High"
            elif "low" in low:
                review = "Low"
            elif "not assessed" in low:
                review = "Not Assessed"

            # Only add rows if we have at least your+expected for each sample
            if y1 is not None and e1 is not None:
                out.append(RCPARow(
                    program=program,
                    participant_id=participant_id,
                    survey_no=survey_no,
                    report_date=report_date,
                    sample_id=left_sample,
                    analyte=analyte,
                    your_result=y1,
                    expected_result=e1,
                    review=review,
                    z_score=z1,
                    aps_score=a1,
                ))

            if y2 is not None and e2 is not None:
                out.append(RCPARow(
                    program=program,
                    participant_id=participant_id,
                    survey_no=survey_no,
                    report_date=report_date,
                    sample_id=right_sample,
                    analyte=analyte,
                    your_result=y2,
                    expected_result=e2,
                    review=review,
                    z_score=z2,
                    aps_score=a2,
                ))
    return out


def extract_all_analytes(full_text: str) -> pd.DataFrame:
    labels = find_analyte_labels(full_text)

    rows = []
    for label in labels:
        rec = parse_single_analyte(full_text, label)
        if rec:
            rows.append(rec)

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["diff_from_mean"] = df["Your Result"] - df["Peer Mean"]
    return df

############################################
# 2. RISK LOGIC (current + historical)
############################################

def flag_bias(last3_diffs: pd.Series) -> bool:
    """
    Bias definition: ≥3 results on the same side of the mean.
    We interpret this as last 3 diffs all >0 or all <0.
    """
    if len(last3_diffs) < 3:
        return False
    return (last3_diffs.gt(0).all()) or (last3_diffs.lt(0).all())


def flag_trend(last3_sdi: pd.Series) -> bool:
    """
    Trend definition: ≥3 results moving in same direction within ±1 SD.
    We'll check:
    - all abs(SDI) <= 1
    - strictly increasing OR strictly decreasing across the last 3 points
    """
    if len(last3_sdi) < 3:
        return False
    if not (last3_sdi.abs() <= 1).all():
        return False
    inc = last3_sdi.iloc[0] < last3_sdi.iloc[1] < last3_sdi.iloc[2]
    dec = last3_sdi.iloc[0] > last3_sdi.iloc[1] > last3_sdi.iloc[2]
    return inc or dec


def base_risk_for_row(row: pd.Series) -> str:
    abs_sdi = abs(row["SDI"]) if pd.notnull(row.get("SDI")) else None

    ts_raw = row.get("Target Score", None)
    ts_num = pd.to_numeric(ts_raw, errors="coerce")
    ts = None if pd.isna(ts_num) else float(ts_num)

    abs_dev = abs(row["%DEV"]) if pd.notnull(row.get("%DEV")) else None

    tea_raw = row.get("TDPA_limit_percent", None)
    tea_num = pd.to_numeric(tea_raw, errors="coerce")
    tea_limit = None if pd.isna(tea_num) else float(tea_num)

    high_hits = []

    if abs_sdi is not None and abs_sdi >= 2:
        high_hits.append("SDI≥2")

    if ts is not None:
        if ts < 40:
            high_hits.append("TS<40")

    if tea_limit is not None and abs_dev is not None:
        if abs_dev > tea_limit:
            high_hits.append(">%TEa")

    if high_hits:
        return "High"

    if (abs_sdi is not None and 1 <= abs_sdi < 2) or (ts is not None and 41 <= ts <= 50):
        return "Moderate"

    return "Low"


def base_risk_for_rcpa(z_score: float, aps_score: float,
                       within_internal_tea: Optional[bool],
                       review: str) -> str:
    """
    RCPA equivalents:
      - z_score ~ SDI
      - APS score > 1 => flagged for review
      - internal TEa (if available)
    """
    abs_z = abs(z_score)

    high_hits = []
    if abs_z >= 2:
        high_hits.append("Z≥2")
    if aps_score > 1.0:
        high_hits.append("APS>1")
    if within_internal_tea is False:
        high_hits.append("Outside_Internal_TEa")
    if review.lower() in {"high", "low"}:
        high_hits.append("Review=High/Low")

    if high_hits:
        return "High"

    # Moderate band
    if (1 <= abs_z < 2) or (0.8 < aps_score <= 1.0):
        return "Moderate"

    return "Low"


def escalate_for_history(hist_df: pd.DataFrame, analyte_name: str, new_row: pd.Series,
                         this_cycle_risk: str) -> tuple[str, bool, bool]:
    """
    Incorporate historical cycles to:
      - upgrade to Critical if recurrent failure (>2 cycles)
      - identify bias/trend over last 3 cycles
    hist_df: full Cycle_History INCLUDING the new row (so last row is current)
    We'll subset hist for this analyte.
    """

    sub = (
        hist_df[hist_df["Analyte"] == analyte_name]
        .sort_values(["Report_Date"])
        .reset_index(drop=True)
    )

    # Bias / Trend flags from last 3 points:
    bias_last3 = flag_bias(sub["diff_from_mean"].tail(3))
    trend_last3 = flag_trend(sub["SDI"].tail(3))

    upgraded_risk = this_cycle_risk

    # Recurrent failure rule:
    # "Recurrent failure (>2 cycles) / Persistent failure -> Critical"
    # We interpret: if this cycle is High AND any High/Critical in the immediately previous cycle,
    # OR if we already had ≥2 High/Critical in the recent past.
    recent = sub["Risk_Category_BaseOnly"].tail(3).tolist()
    # Count how many High-or-worse in the last 3 (including now)
    high_like_count = sum(r in ["High", "Critical"] for r in recent)
    if high_like_count >= 2 and this_cycle_risk in ["High", "Critical"]:
        upgraded_risk = "Critical"

    return upgraded_risk, bias_last3, trend_last3


def build_comment_and_action(final_risk: str,
                             bias_flagged: bool,
                             trend_flagged: bool) -> tuple[str, str]:
    """
    Gives:
    - Interpretation / Comments (goes into Result Summary, auditable narrative)
    - Required Action (for escalation sheet)
    """
    if final_risk == "Critical":
        action = "Escalate to LM; notify QA"
        comment = "Persistent failure (≥2 cycles). Immediate escalation required."
        return comment, action

    if final_risk == "High":
        action = "Investigate immediately; complete RCA and alert LM"
        comment = "Outside TEa / SDI ≥2 or low target score. Immediate RCA required."
        return comment, action

    # Moderate
    if final_risk == "Moderate":
        action = "Monitor next EQA cycles; review calibration"
        if bias_flagged and trend_flagged:
            comment = "Developing bias and trend (≥3 results consistent direction within ±1 SD). Monitor closely."
        elif bias_flagged:
            comment = "Developing bias (≥3 results on same side of mean). Monitor next cycle."
        elif trend_flagged:
            comment = "Developing trend (drift over ≥3 cycles within ±1 SD). Monitor next cycle."
        else:
            comment = "Slight bias/trend or borderline target score (41–50). Monitor."
        return comment, action

    # Low
    action = "File record; continue routine QC"
    comment = "Acceptable: within TEa, SDI<2, TS>50. No significant bias/trend."
    return comment, action


############################################
# 3. EXCEL UPDATE
############################################
def norm_analyte(s: str) -> str:
    if s is None:
        return ""
    s = str(s)

    # common hidden junk from PDFs/Excel
    s = s.replace("\xa0", " ")      # non-breaking space
    s = s.replace("Â", "")          # common PDF artefact before degree sign

    # normalise degree variants
    s = s.replace("º", "°")

    # collapse whitespace
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = s.replace("uric acid (urate), mg/dl", "uric acid (urate), mmol/l")
    return s


def load_internal_tea_map(path: Path) -> Dict[str, Optional[float]]:
    """
    Expects columns:
      - Analyte
      - Internal_TEa
    Values can be numeric or 'N/A'.
    Returns dict: normalized_analyte -> TEa (float) or None
    """
    path = Path(path)

    if path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        # Handles CSV or TSV; if your file is tab-delimited this will still work
        df = pd.read_csv(path, sep=None, engine="python")

    # Column normalization (so minor header spelling/case doesn't break it)
    cols = {c: c.strip() for c in df.columns}
    df = df.rename(columns=cols)

    if "Analyte" not in df.columns or "Internal_TEa" not in df.columns:
        raise ValueError(f"Internal TEa file must have columns 'Analyte' and 'Internal_TEa'. Found: {list(df.columns)}")

    tea_map: Dict[str, Optional[float]] = {}

    for _, r in df.iterrows():
        a = _norm_analyte_name(r["Analyte"])
        v = r["Internal_TEa"]

        if pd.isna(v):
            tea_map[a] = None
            continue

        vs = str(v).strip()
        if vs == "" or vs.lower() in {"n/a", "na", "none", "null"}:
            tea_map[a] = None
            continue

        # If someone ever uses commas again, handle it safely
        vs = vs.replace(",", ".")
        tea_map[a] = float(vs)

    return tea_map


def update_latest_cycle_sheet(wb: Workbook):
    """
    Create/refresh a sheet called 'Latest_Cycle' which shows ONLY
    the most recent cycle (max Report_Date) from Cycle_History.
    """
    if "Cycle_History" not in wb.sheetnames:
        return

    hist_ws = wb["Cycle_History"]
    hist_df = ws_to_dataframe(hist_ws)

    if hist_df.empty:
        return

    # Ensure dates are proper datetime
    hist_df["Report_Date"] = pd.to_datetime(hist_df["Report_Date"], errors="coerce")
    hist_df = hist_df.dropna(subset=["Report_Date"])

    if hist_df.empty:
        return

    latest_date = hist_df["Report_Date"].max()
    latest_df = hist_df[hist_df["Report_Date"] == latest_date].copy()

    # (Optional) sort nicely for reviewer
    latest_df = latest_df.sort_values(["Risk_Category_Final", "Analyte"], ascending=[False, True])

    # Create or clear sheet
    if "Latest_Cycle" in wb.sheetnames:
        ws = wb["Latest_Cycle"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Latest_Cycle")

    # Write header
    headers = list(latest_df.columns)
    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i).value = h

    # Write rows
    for r_i, row in enumerate(latest_df.itertuples(index=False), start=2):
        for c_i, val in enumerate(row, start=1):
            ws.cell(row=r_i, column=c_i).value = val

def ensure_cycle_history_sheet(wb: Workbook):
    """Create Cycle_History sheet with headers if it doesn't exist yet."""
    if "Cycle_History" not in wb.sheetnames:
        ws = wb.create_sheet("Cycle_History")
        headers = [
            "Cycle_No",
            "Sample_No",
            "Report_Date",
            "Analyte",
            "Peer_Mean",
            "Your_Result",
            "%DEV",
            "SDI",
            "Target_Score",  # RIQAS uses this
            "TEa_or_TDPA(%)",  # RIQAS uses this
            "APS",  # ✅ NEW (RCPA)
            "Review",  # ✅ NEW (RCPA)
            "Internal_TEa",
            "diff_from_mean",
            "Risk_Category_BaseOnly",
            "Risk_Category_Final",
            "Bias_Flag_Last3",
            "Trend_Flag_Last3",
            "Within_Internal_TEa?",
            "Required_Action",
            "Comment",
        ]

        for col_i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_i).value = header

def ensure_cycle_history_columns(wb: Workbook, include_rcpa_cols: bool = False):
    """
    If Cycle_History already exists but is missing the new columns,
    insert them at the correct positions.
    """
    if "Cycle_History" not in wb.sheetnames:
        return

    ws = wb["Cycle_History"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    def find_col(name: str):
        return headers.index(name) + 1 if name in headers else None

    # --- Insert APS + Review ONLY for RCPA ---
    if include_rcpa_cols:
        # --- Insert APS after TEa_or_TDPA(%) ---
        if "APS" not in headers:
            tea_col = find_col("TEa_or_TDPA(%)")
            if tea_col is None:
                raise ValueError("Cycle_History missing 'TEa_or_TDPA(%)'")
            ws.insert_cols(tea_col + 1)
            ws.cell(row=1, column=tea_col + 1).value = "APS"

            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h is not None else "" for h in headers]

        # --- Insert Review after APS ---
        if "Review" not in headers:
            aps_col = find_col("APS")
            if aps_col is None:
                raise ValueError("Cycle_History missing 'APS' (cannot insert Review)")
            ws.insert_cols(aps_col + 1)
            ws.cell(row=1, column=aps_col + 1).value = "Review"

            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h is not None else "" for h in headers]

    # --- Insert Internal_TEa after Review ---
    if "Internal_TEa" not in headers:
        review_col = find_col("Review")
        if review_col is None:
            # if Review not present for some reason, fall back to TEa_or_TDPA(%)
            review_col = find_col("TEa_or_TDPA(%)")
        if review_col is None:
            raise ValueError("Cycle_History missing 'TEa_or_TDPA(%)'/'Review'")
        ws.insert_cols(review_col + 1)
        ws.cell(row=1, column=review_col + 1).value = "Internal_TEa"

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in headers]

    # --- Insert Within_Internal_TEa? after Trend_Flag_Last3 ---
    if "Within_Internal_TEa?" not in headers:
        trend_col = find_col("Trend_Flag_Last3")
        if trend_col is None:
            raise ValueError("Cycle_History missing 'Trend_Flag_Last3'")
        ws.insert_cols(trend_col + 1)
        ws.cell(row=1, column=trend_col + 1).value = "Within_Internal_TEa?"

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in headers]


def extract_rcpa_auto_diff(full_text: str) -> pd.DataFrame:
    """
    Extract RCPA Automated Differential (Option 1) analytes.
    We match each analyte block like:

        Neutrophils
        47.80
        48.40
        Within APS
        -0.8
        -0.1
    """

    analytes = [
        "White cell count",
        "Neutrophils",
        "Lymphocytes",
        "Monocytes",
        "Eosinophils",
        "Basophils",
        "Immature Granulocytes",
    ]

    rows = []

    for a in analytes:
        # analyte + your + expected + review + zscore + aps
        pat = re.compile(
            rf"{re.escape(a)}\s+([-\d.]+)\s+([-\d.]+)\s+(Within APS|Review Required|Outside APS|Within\s+APS)\s+([-\d.]+)\s+([-\d.]+)",
            re.IGNORECASE
        )
        m = pat.search(full_text)
        if not m:
            continue

        your_result = float(m.group(1))
        expected = float(m.group(2))
        review = m.group(3).strip()
        zscore = float(m.group(4))
        aps = float(m.group(5))  # keep it, but DON'T treat as %TEa

        # %DEV from expected
        pct_dev = None
        if expected != 0:
            pct_dev = ((your_result - expected) / expected) * 100.0

        rows.append({
            "AnalyteRaw": a,
            "Peer Mean": expected,          # expected = peer/target in this report
            "Your Result": your_result,
            "%DEV": pct_dev,
            "SDI": zscore,                  # treat Z-score like SDI
            "Target Score": None,           # RCPA doesn't provide TS
            "TDPA_limit_percent": None,     # RCPA APS is NOT %TEa, so don't use it
            "RCPA_APS": aps,
            "RCPA_Review": review,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["diff_from_mean"] = df["Your Result"] - df["Peer Mean"]
    return df

def ws_to_dataframe(ws) -> pd.DataFrame:
    """Helper: read an openpyxl worksheet into a DataFrame (header row is row 1)."""
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    df = pd.DataFrame(data[1:], columns=data[0])
    # Drop completely empty rows
    df = df.dropna(how="all")
    return df

def backfill_internal_tea_in_cycle_history(wb: Workbook, internal_tea_map: dict):
    """
    For existing Cycle_History rows, fill Internal_TEa and Within_Internal_TEa?
    where they are blank, using the current internal_tea_map.
    """
    if "Cycle_History" not in wb.sheetnames:
        return

    ws = wb["Cycle_History"]

    # Read sheet into DF
    hist_df = ws_to_dataframe(ws)
    if hist_df.empty:
        return

    # Make sure required columns exist in DF (in case header row exists but DF is older)
    if "Analyte" not in hist_df.columns or "%DEV" not in hist_df.columns:
        return
    if "Internal_TEa" not in hist_df.columns:
        hist_df["Internal_TEa"] = None
    if "Within_Internal_TEa?" not in hist_df.columns:
        hist_df["Within_Internal_TEa?"] = None

    # Normalize analyte + map TEa
    hist_df["_tea_key"] = hist_df["Analyte"].apply(norm_analyte)
    mapped = hist_df["_tea_key"].map(internal_tea_map)

    # Only fill Internal_TEa if blank
    internal_blank = hist_df["Internal_TEa"].isna() | (hist_df["Internal_TEa"].astype(str).str.strip() == "")
    hist_df.loc[internal_blank, "Internal_TEa"] = mapped[internal_blank]

    # Recompute Within_Internal_TEa? where blank AND TEa exists
    def within_internal_hist(r):
        try:
            tea = r["Internal_TEa"]
            dev = r["%DEV"]
            if pd.isna(tea) or str(tea).strip() == "":
                return ""
            if pd.isna(dev) or str(dev).strip() == "":
                return ""
            return "Yes" if abs(float(dev)) <= float(tea) else "No"
        except Exception:
            return ""

    within_blank = hist_df["Within_Internal_TEa?"].isna() | (hist_df["Within_Internal_TEa?"].astype(str).str.strip() == "")
    hist_df.loc[within_blank, "Within_Internal_TEa?"] = hist_df.loc[within_blank].apply(within_internal_hist, axis=1)

    hist_df = hist_df.drop(columns=["_tea_key"], errors="ignore")

    # Rewrite the sheet (keep current header order exactly as in Excel)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    # Clear everything below header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Append rows in the exact header order
    for _, r in hist_df.iterrows():
        row_out = [r.get(h, None) for h in headers]
        ws.append(row_out)


def append_df_to_worksheet(ws, df: pd.DataFrame):
    """
    Append df rows into ws by matching the worksheet header names.
    Prevents column-shift bugs when columns are inserted/reordered.
    """
    # sheet headers (row 1)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    for _, r in df.iterrows():
        out_row = []
        for h in headers:
            out_row.append(r.get(h, None))
        ws.append(out_row)



def fill_header_information_sheet(wb: Workbook, meta: dict):
    """
    Populate 'Header Information' sheet:
    A1: 'Laboratory Name' -> put value in B1, etc.
    We'll just write into column B for the known rows.
    """
    if "Header Information" not in wb.sheetnames:
        return
    ws = wb["Header Information"]

    mapping = {
        "Laboratory Name": meta.get("lab_name"),
        "Instrument/Method": meta.get("instrument_name"),
        "Analyte / Panel": meta.get("panel_name"),
        "Cycle / Sample No.": f"Cycle {meta.get('cycle_no')} / Sample {meta.get('sample_no')}",
        "Date of Report": meta.get("report_date").strftime("%d/%m/%Y") if meta.get("report_date") else None,
        "EQA Scheme": "RIQAS (Randox International Quality Assessment Scheme)",  # already present
    }

    # loop rows, put in column B (col 2)
    for row in range(1, 20):
        label = ws.cell(row=row, column=1).value
        if label in mapping and mapping[label] is not None:
            ws.cell(row=row, column=2).value = mapping[label]


def map_parameter_name(analyte_raw: str) -> str:
    """
    Convert the RIQAS analyte label into the row label used in 'Result Summary'.
    This mapping can be expanded as needed.
    """
    mapping_table = {
        "Haemoglobin, g/dl": "Haemoglobin (g/dL)",
        "Haematocrit (HCT), L/L": "Haematocrit (L/L)",
        "MCH, pg": "MCH (pg)",
        "MCHC, g/dl": "MCHC (g/dL)",
        "MCV, fL": "MCV (fL)",
        "Mean Platelet Volume, fL": "MPV (fL)",
        "Plateletcrit, %": "Plateletcrit (%)",
        "Platelets (Impedance Count), X 10 9/L": "Platelets (x10^9/L)",
        "RBC (Impedance Count), X 10 12/L": "RBC (x10^12/L)",
        "Red Cell Dist. Width CV, %": "RDW-CV (%)",
        "WBC (Optical Count), X 10 9/L": "WBC (x10^9/L)",
    }
    return mapping_table.get(analyte_raw, analyte_raw)


def update_result_summary_sheet(wb: Workbook,
                                current_cycle_rows: pd.DataFrame,
                                final_eval: pd.DataFrame):
    """
    'Result Summary' sheet has this header:
        A: Parameter
        B: Peer Group Mean
        C: Your Result
        D: %Deviation
        E: SDI
        F: Target Score
        G: Acceptable Limit (TEa or RIQAS)
        H: Acceptable? (Y/N)
        I: Interpretation / Comments

    We'll overwrite rows 2..N with the latest cycle values.
    We'll also inject a standardised comment aligned to SANAS language.
    """
    if "Result Summary" not in wb.sheetnames:
        return

    ws = wb["Result Summary"]

    # First: clear old numeric cells in rows 2-50 for cols B..I
    for row in range(2, 60):
        for col in range(2, 10):  # B..I
            ws.cell(row=row, column=col).value = None

    # Build a lookup of risk/action/comment from final_eval
    # final_eval columns:
    #   Analyte, Risk_Category_Final, Comment, Required_Action
    eval_lookup = (
        final_eval.set_index("Parameter_Name")
                  [["Risk_Category_Final", "Comment", "Required_Action", "TDPA_limit_percent"]]
        .to_dict(orient="index")
    )

    # Now walk the sheet and if Parameter cell matches, fill.
    for row in range(2, 200):
        param_cell_val = ws.cell(row=row, column=1).value
        if not param_cell_val:
            continue

        # find matching record in current_cycle_rows
        subset = current_cycle_rows[current_cycle_rows["Parameter_Name"] == param_cell_val]
        if subset.empty:
            continue
        rec = subset.iloc[0]

        # acceptable? logic for column H:
        # Acceptable if:
        #   SDI < 2
        #   Target Score > 50
        #   abs(%DEV) <= TEa
        # acceptable? logic for column H (hardened: TS/TEa may be strings)
        abs_sdi = abs(float(rec["SDI"])) if pd.notnull(rec.get("SDI")) and str(rec.get("SDI")).strip() != "" else None

        ts_raw = rec.get("Target Score", None)
        ts_num = pd.to_numeric(ts_raw, errors="coerce")
        ts = None if pd.isna(ts_num) else float(ts_num)

        dev_raw = rec.get("%DEV", None)
        dev_num = pd.to_numeric(dev_raw, errors="coerce")
        abs_dev = None if pd.isna(dev_num) else abs(float(dev_num))

        # Prefer Internal_TEa; else fall back to TDPA_limit_percent
        tea_raw = rec.get("Internal_TEa", None)
        tea_num = pd.to_numeric(tea_raw, errors="coerce")
        if pd.isna(tea_num):
            tea_raw = rec.get("TDPA_limit_percent", None)
            tea_num = pd.to_numeric(tea_raw, errors="coerce")
        tea_lim = None if pd.isna(tea_num) else float(tea_num)

        acceptable = (
                (abs_sdi is not None and abs_sdi < 2) and
                ((ts is None) or (ts > 50)) and
                ((tea_lim is None) or (abs_dev is not None and abs_dev <= tea_lim))
        )

        acceptable_flag = "Y" if acceptable else "N"

        # Pull risk-derived narrative
        risk_info = eval_lookup.get(param_cell_val, {})
        narrative_comment = risk_info.get("Comment", "")
        tea_display = rec["TDPA_limit_percent"]

        ws.cell(row=row, column=2).value = float(rec["Peer Mean"])
        ws.cell(row=row, column=3).value = float(rec["Your Result"])
        ws.cell(row=row, column=4).value = float(rec["%DEV"]) if pd.notnull(rec["%DEV"]) else None
        ws.cell(row=row, column=5).value = float(rec["SDI"])
        ws.cell(row=row, column=6).value = float(rec["Target Score"])
        ws.cell(row=row, column=7).value = tea_display
        ws.cell(row=row, column=8).value = acceptable_flag
        ws.cell(row=row, column=9).value = narrative_comment


def update_cycle_history_sheet(wb: Workbook,
                               meta: dict,
                               enriched_rows: pd.DataFrame):
    """
    Append new rows (the latest cycle) into Cycle_History.
    Cycle_History is the auditable longitudinal evidence for SANAS:
      - shows consecutive bias, trend, recurrence
      - shows final risk and action
      - supports "Critical" escalation
    """
    ws = wb["Cycle_History"]

    append_df_to_worksheet(ws, enriched_rows)


############################################
# 4. MAIN ORCHESTRATION
############################################
RCPA_HAEME_ANALYTES = [
    "White cell count",
    "Neutrophils",
    "Lymphocytes",
    "Monocytes",
    "Eosinophils",
    "Basophils",
    "Immature Granulocytes",
]

def _to_float(x: str):
    try:
        return float(x)
    except Exception:
        return None

def parse_rcpa_automated_diff_blocks(full_text: str) -> pd.DataFrame:
    """
    Robust RCPA Automated Differential parser:
    For each analyte heading, take the next numeric lines in the block:
      - Your Result (first number)
      - Expected Result (second number)
      - Z-score (first "small" number after 'Within APS' if present; else next number)
    APS / Target-score-like fields are often mangled by pdfminer; we treat them as optional.
    """

    txt = normalize_pdf_text(full_text)
    rows = []

    for analyte in RCPA_HAEME_ANALYTES:
        m = re.search(rf"^{re.escape(analyte)}\s*$", txt, flags=re.IGNORECASE | re.MULTILINE)
        if not m:
            continue

        block = txt[m.end(): m.end() + 800]  # window after analyte
        lines = [ln.strip() for ln in block.splitlines() if ln.strip()]

        # collect numeric lines in the block (first ~10 is usually enough)
        nums = []
        for ln in lines[:20]:
            # keep only plain numbers like 17.66, -0.6 etc
            if re.fullmatch(r"-?\d+(\.\d+)?", ln):
                nums.append(float(ln))

        # Need at least Your + Expected
        if len(nums) < 2:
            continue

        your = nums[0]
        expected = nums[1]

        # try to find Z-score: often the next numbers include z-score(s)
        # Heuristic: choose the next value with abs <= 6 (typical z-score range)
        z = None
        for cand in nums[2:]:
            if cand is None:
                continue
            if abs(cand) <= 6:
                z = cand
                break

        # %DEV from your vs expected
        dev = None
        if expected not in (None, 0):
            dev = ((your - expected) / expected) * 100.0

        rows.append({
            "AnalyteRaw": analyte,
            "Peer Mean": expected,          # treat expected as peer/assigned for now
            "Your Result": your,
            "%DEV": dev,
            "SDI": z,                       # z-score ~ SDI for your risk logic
            "Target Score": None,           # not reliably available -> None
            "TDPA_limit_percent": None,     # not in RCPA; we'll use Internal_TEa instead
        })

    return pd.DataFrame(rows)

def process_riqas_pdf_into_workbook(pdf_path: str, xlsx_path: str, out_path: Optional[str] = None,
                                    internal_tea_map: Optional[dict] = None):

    """
    Full pipeline:
      - Read PDF
      - Extract analyte performance
      - Load workbook
      - Update Header Information
      - Update Result Summary (this cycle)
      - Update/append Cycle_History (all cycles, classified with history-aware risk)
      - Save workbook (in-place or to new file)
    """
    pdf_path = Path(pdf_path)
    xlsx_path = Path(xlsx_path)
    out_path = Path(out_path) if out_path else xlsx_path

    # --- parse PDF ---
    full_text = read_pdf_text(pdf_path)
    full_text = normalize_pdf_text(full_text)

    # ==========================================================
    # RCPA branch (Option B): ONLY update Cycle_History and EXIT
    # ==========================================================
    if is_rcpa_report(full_text):
        internal_tea_map = internal_tea_map or {}

        # Load or create workbook
        wb = load_workbook(xlsx_path) if xlsx_path.exists() else Workbook()

        # Ensure Cycle_History exists and is valid
        ensure_cycle_history_sheet(wb)
        ensure_cycle_history_columns(wb, include_rcpa_cols=True)
        backfill_internal_tea_in_cycle_history(wb, internal_tea_map)

        # Convert existing history to DataFrame (for risk calc if needed)
        hist_ws = wb["Cycle_History"]
        hist_df_existing = (
            ws_to_dataframe(hist_ws)
            if hist_ws.max_row > 1
            else pd.DataFrame(columns=[
                "Cycle_No",
                "Sample_No",
                "Report_Date",
                "Analyte",
                "Peer_Mean",
                "Your_Result",
                "%DEV",
                "SDI",
                "Target_Score",
                "TEa_or_TDPA(%)",
                "Internal_TEa",
                "diff_from_mean",
                "Risk_Category_BaseOnly",
                "Risk_Category_Final",
                "Bias_Flag_Last3",
                "Trend_Flag_Last3",
                "Within_Internal_TEa?",
                "Required_Action",
                "Comment",
            ])
        )

        hist_df_existing = coerce_cycle_history_types(hist_df_existing)

        # >>> IMPORTANT <<<
        # RCPA reports do NOT update Header Information,
        # Result Summary, or Latest_Cycle

        # ----------------------------------------------------------
        # RCPA metadata (so Report_Date is ALWAYS populated)
        # ----------------------------------------------------------
        rcpa_report_date = parse_rcpa_issue_date(full_text)     # Issue date in PDF
        rcpa_survey_no = parse_rcpa_survey_no(full_text)        # Survey number (int)

        sm = re.search(r"(HA-[A-Z0-9]+-\d{2}-\d{2})", full_text)
        rcpa_sample_id = sm.group(1) if sm else "RCPA"

        report_dt = pd.to_datetime(rcpa_report_date, errors="coerce")

        # ----------------------------------------------------------
        # Parse RCPA analytes
        # Prefer robust block parser; fallback to Summary table
        # ----------------------------------------------------------
        # Prefer Summary table because it includes BOTH samples explicitly
        rcpa_rows = parse_rcpa_summary_of_performance(full_text)
        df_rcpa = pd.DataFrame()

        if rcpa_rows:
            tmp = []
            for rr in rcpa_rows:
                dev = None if rr.expected_result == 0 else ((
                                                                        rr.your_result - rr.expected_result) / rr.expected_result) * 100.0
                tmp.append({
                    "Survey_No": rr.survey_no,
                    "Sample_ID": rr.sample_id,
                    "AnalyteRaw": rr.analyte.strip(),
                    "Peer Mean": rr.expected_result,
                    "Your Result": rr.your_result,
                    "%DEV": dev,
                    "SDI": rr.z_score,
                    "Target Score": None,
                    "TDPA_limit_percent": None,
                    "diff_from_mean": rr.your_result - rr.expected_result,
                    "RCPA_Review": rr.review,
                    "RCPA_APS": rr.aps_score,
                })
            df_rcpa = pd.DataFrame(tmp)

        # fallback to block parser ONLY if summary table fails
        if df_rcpa is None or df_rcpa.empty:
            df_rcpa = parse_rcpa_automated_diff_blocks(full_text)
            # block parser doesn't know sample per row -> tag as UNKNOWN
            if df_rcpa is not None and not df_rcpa.empty:
                df_rcpa["Survey_No"] = rcpa_survey_no
                df_rcpa["Sample_ID"] = "UNKNOWN"
                df_rcpa["RCPA_APS"] = df_rcpa.get("RCPA_APS", None)
                df_rcpa["RCPA_Review"] = df_rcpa.get("RCPA_Review", "")

            # --- RCPA: normalise column names (prevents KeyError on different report types) ---
            # Some RCPA parsers produce snake_case, some produce human names.
            rename_map = {
                "your_result": "Your Result",
                "expected_result": "Peer Mean",
                "peer_mean": "Peer Mean",
                "mean": "Peer Mean",
                "expected": "Peer Mean",
            }
            df_rcpa = df_rcpa.rename(columns=rename_map)

            required_cols = {"Your Result", "Peer Mean"}
            missing = required_cols - set(df_rcpa.columns)
            if missing:
                raise ValueError(
                    f"RCPA parse produced no usable numeric table for this program. "
                    f"Missing columns: {sorted(missing)}. Found: {list(df_rcpa.columns)}"
                )

        # Ensure diff_from_mean exists (block parser doesn't always set it)
        if "diff_from_mean" not in df_rcpa.columns:
            df_rcpa["diff_from_mean"] = df_rcpa["Your Result"] - df_rcpa["Peer Mean"]

        # ----------------------------------------------------------
        # Dedupe keys: (Report_Date, Sample_No, Analyte)
        # ----------------------------------------------------------
        existing_keys = set()
        if not hist_df_existing.empty:
            for _, ex in hist_df_existing.iterrows():
                rd = ex.get("Report_Date")
                sn = ex.get("Sample_No")
                an = ex.get("Analyte")
                if pd.isna(rd) or pd.isna(sn) or not an:
                    continue
                existing_keys.add((str(pd.to_datetime(rd).date()), str(sn), str(an).strip().lower()))

        enriched_rows_for_history = []

        # Use report date from RCPA header if available (Report Issue Date)
        report_date = parse_rcpa_report_issue_date(full_text)
        report_dt = pd.to_datetime(report_date, errors="coerce")

        # Try to detect the 2 sample IDs from the report (so each sample becomes separate rows)
        sample_ids = []
        for l in full_text.splitlines():
            m = re.search(r"Sample:\s*(HA-[A-Z0-9]+-\d{2}-\d{2})", l, re.IGNORECASE)
            if m:
                sample_ids.append(m.group(1))
        # de-dupe preserve order
        seen = set()
        sample_ids = [s for s in sample_ids if not (s in seen or seen.add(s))]

        # Fallback sample IDs if parsing failed
        left_sample = sample_ids[0] if len(sample_ids) >= 1 else "UNKNOWN_LEFT"
        right_sample = sample_ids[1] if len(sample_ids) >= 2 else "UNKNOWN_RIGHT"

        # Build existing keys for dedupe
        existing_keys = set()
        if not hist_df_existing.empty:
            for _, hx in hist_df_existing.iterrows():
                rd = hx.get("Report_Date")
                sn = hx.get("Sample_No")
                an = hx.get("Analyte")
                if pd.isna(rd) or pd.isna(sn) or not an:
                    continue
                existing_keys.add((str(pd.to_datetime(rd).date()), str(sn), str(an).strip().lower()))

        for _, r in df_rcpa.iterrows():
            analyte_name = str(r.get("AnalyteRaw", "")).strip()
            if not analyte_name:
                continue

            expected = float(r.get("Peer Mean")) if pd.notna(r.get("Peer Mean")) else None
            your = float(r.get("Your Result")) if pd.notna(r.get("Your Result")) else None
            z = float(r.get("SDI")) if pd.notna(r.get("SDI")) else None
            aps_val = float(r.get("RCPA_APS")) if pd.notna(r.get("RCPA_APS")) else None
            review_txt = str(r.get("RCPA_Review", "")).strip()

            if expected is None or your is None:
                continue

            # %DEV
            dev = None
            if expected != 0:
                dev = ((your - expected) / expected) * 100.0

            # Internal TEa lookup (optional)
            tea_key = _norm_analyte_name(analyte_name)
            internal_tea = (internal_tea_map or {}).get(tea_key)

            within_internal = None
            if internal_tea is not None and dev is not None:
                within_internal = (abs(dev) <= float(internal_tea))

            # sample id from df_rcpa if present; else infer left/right by "Sample_Which"
            sample_no = r.get("Sample_ID")
            if pd.isna(sample_no) or not str(sample_no).strip():
                which = str(r.get("Sample_Which", "")).strip().lower()
                if which == "right":
                    sample_no = right_sample
                else:
                    sample_no = left_sample
            sample_no = str(sample_no).strip()

            dedupe_key = (str(report_dt.date()) if pd.notna(report_dt) else "", sample_no, analyte_name.lower())
            if dedupe_key in existing_keys:
                continue

            new_point = {
                "Cycle_No": int(r.get("Survey_No")) if pd.notna(r.get("Survey_No")) else None,
                "Sample_No": sample_no,
                "Report_Date": report_dt,
                "Analyte": analyte_name,
                "APS": aps_val,
                "Review": review_txt,
                "Peer_Mean": expected,
                "Your_Result": your,
                "%DEV": dev,
                "SDI": z,

                "Target_Score": None,
                "TEa_or_TDPA(%)": None,

                "Internal_TEa": internal_tea,
                "diff_from_mean": (your - expected),
                "Risk_Category_BaseOnly": None,
                "Risk_Category_Final": None,
                "Bias_Flag_Last3": None,
                "Trend_Flag_Last3": None,
                "Within_Internal_TEa?": (
                    "Yes" if within_internal is True else "No" if within_internal is False else ""),
                "Required_Action": None,
                "Comment": None,
            }

            new_point_df = pd.DataFrame([new_point])

            # historical subset for this analyte
            sub_hist = pd.DataFrame()
            if not hist_df_existing.empty and "Analyte" in hist_df_existing.columns:
                sub_hist = hist_df_existing[
                    hist_df_existing["Analyte"].astype(str).str.lower() == analyte_name.lower()
                    ].copy()

            simulated_hist = new_point_df.copy() if sub_hist.empty else pd.concat([sub_hist, new_point_df],
                                                                                  ignore_index=True, sort=False)
            simulated_hist["Report_Date"] = pd.to_datetime(simulated_hist["Report_Date"], errors="coerce")
            simulated_hist = simulated_hist.sort_values(["Report_Date"]).reset_index(drop=True)

            # base risk (RCPA): we have z and (optionally) internal TEa; APS is treated like “score”
            base_risk_label = base_risk_for_rcpa(
                z_score=(z if z is not None else 0.0),
                aps_score=(aps_val if aps_val is not None else 0.0),
                within_internal_tea=within_internal,
                review=review_txt,
            )

            simulated_hist.loc[simulated_hist.index[-1], "Risk_Category_BaseOnly"] = base_risk_label

            final_risk_label, bias_last3, trend_last3 = escalate_for_history(
                simulated_hist,
                analyte_name,
                pd.Series({"SDI": z}),
                base_risk_label,
            )

            comment_txt, action_txt = build_comment_and_action(final_risk_label, bias_last3, trend_last3)

            last = simulated_hist.index[-1]
            simulated_hist.loc[last, "Risk_Category_Final"] = final_risk_label
            simulated_hist.loc[last, "Bias_Flag_Last3"] = bool(bias_last3)
            simulated_hist.loc[last, "Trend_Flag_Last3"] = bool(trend_last3)
            simulated_hist.loc[last, "Required_Action"] = action_txt
            simulated_hist.loc[last, "Comment"] = f"{comment_txt} (Review={review_txt}, APS={aps_val})"

            enriched_rows_for_history.append(simulated_hist.iloc[[-1]])

        # Nothing new? still refresh Latest_Cycle and save
        if not enriched_rows_for_history:
            update_latest_cycle_sheet(wb)
            wb.save(out_path)
            return

        enriched_df = pd.concat(enriched_rows_for_history, ignore_index=True)

        # Append to Cycle_History
        update_cycle_history_sheet(wb, meta={}, enriched_rows=enriched_df[
            [
                "Cycle_No","Sample_No","Report_Date","Analyte","Peer_Mean","Your_Result",
                "%DEV","SDI","Target_Score","TEa_or_TDPA(%)","APS","Review","Internal_TEa","diff_from_mean",
                "Risk_Category_BaseOnly","Risk_Category_Final","Bias_Flag_Last3","Trend_Flag_Last3",
                "Within_Internal_TEa?","Required_Action","Comment"
            ]
        ])

        update_latest_cycle_sheet(wb)
        wb.save(out_path)
        return

    meta = parse_metadata(full_text)

    t = full_text.lower()
    is_rcpa = ("rcpa" in t) or ("australian" in t and "quality" in t)
    is_auto_diff = ("automated differential" in t) or ("auto differential" in t) or (
                "neutrophils" in t and "lymphocytes" in t)

    if is_rcpa and is_auto_diff:
        df = parse_rcpa_automated_diff_blocks(full_text)
        if df is None or df.empty:
            df = extract_all_analytes(full_text)  # fallback
    else:
        df = extract_all_analytes(full_text)

    if df.empty:
        raise ValueError("No analytes extracted from this PDF (layout/label detection failed).")
    print("  meta:", meta)
    print("  analytes found:", len(df))

    # Attach metadata columns now (cycle/date/sample etc)
    df["Cycle_No"] = meta["cycle_no"]
    df["Sample_No"] = meta["sample_no"]
    df["Report_Date"] = meta["report_date"]

    # ✅ FORCE pandas Timestamp (critical for comparisons & sorting)
    df["Report_Date"] = pd.to_datetime(df["Report_Date"], errors="coerce")

    internal_tea_map = internal_tea_map or {}

    df["_tea_key"] = df["AnalyteRaw"].apply(norm_analyte)
    df["Internal_TEa"] = df["_tea_key"].map(internal_tea_map)

    def within_internal(row):
        if pd.isna(row["Internal_TEa"]) or pd.isna(row["%DEV"]):
            return ""
        return "Yes" if abs(float(row["%DEV"])) <= float(row["Internal_TEa"]) else "No"

    df["Within_Internal_TEa?"] = df.apply(within_internal, axis=1)
    df = df.drop(columns=["_tea_key"])

    missing = df[df["Internal_TEa"].isna()]["AnalyteRaw"].tolist()
    if missing:
        print("⚠️ No Internal_TEa match for:", missing)
    else:
        print("✅ Internal_TEa matched for all analytes in this PDF.")

    # Rename for workbook logic
    df["Parameter_Name"] = df["AnalyteRaw"]

    # --- open workbook / init sheets ---
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    ensure_cycle_history_sheet(wb)
    ensure_cycle_history_columns(wb, include_rcpa_cols=False)
    # ✅ backfill old rows that existed before these columns were added
    backfill_internal_tea_in_cycle_history(wb, internal_tea_map)

    # We'll convert Cycle_History (existing) to df_hist for historical risk calc
    hist_ws = wb["Cycle_History"]
    hist_df_existing = ws_to_dataframe(hist_ws) if hist_ws.max_row > 1 else pd.DataFrame(columns=[
        "Cycle_No",
        "Sample_No",
        "Report_Date",
        "Analyte",
        "Peer_Mean",
        "Your_Result",
        "%DEV",
        "SDI",
        "Target_Score",
        "TEa_or_TDPA(%)",
        "Internal_TEa",  # ✅
        "diff_from_mean",
        "Risk_Category_BaseOnly",
        "Risk_Category_Final",
        "Bias_Flag_Last3",
        "Trend_Flag_Last3",
        "Within_Internal_TEa?",  # ✅
        "Required_Action",
        "Comment",
    ])
    hist_df_existing = coerce_cycle_history_types(hist_df_existing)

    if not hist_df_existing.empty:
        hist_df_existing["Report_Date"] = pd.to_datetime(
            hist_df_existing["Report_Date"], errors="coerce"
        )
    # --- STEP 4: build set of existing cycle/analyte keys (deduplication) ---
    existing_keys = set()

    if not hist_df_existing.empty:
        for _, r in hist_df_existing.iterrows():
            c = r.get("Cycle_No")
            s = r.get("Sample_No")
            a = r.get("Analyte")
            if pd.isna(c) or pd.isna(s) or not a:
                continue
            existing_keys.add((int(c), int(s), str(a).strip()))

    # --- build risk per analyte, including historical escalation ---
    enriched_rows_for_history = []
    summary_eval_rows = []

    for _, row in df.iterrows():
        analyte_name = row["Parameter_Name"]

        key = (row["Cycle_No"], row["Sample_No"], analyte_name)
        if key in existing_keys:
            continue

        # Build a temp frame with existing + this new measurement for THIS analyte
        sub_hist = hist_df_existing[hist_df_existing["Analyte"] == analyte_name].copy()

        # --- STEP 5: ensure historical base risk is populated ---
        if not sub_hist.empty:
            for idx_hist, r in sub_hist.iterrows():
                if pd.isna(r.get("Risk_Category_BaseOnly")):
                    # reconstruct a pseudo-row compatible with base_risk_for_row
                    pseudo = pd.Series({
                        "SDI": r["SDI"],
                        "Target Score": r["Target_Score"],
                        "%DEV": r["%DEV"],
                        "TDPA_limit_percent": r["TEa_or_TDPA(%)"],
                    })
                    sub_hist.at[idx_hist, "Risk_Category_BaseOnly"] = base_risk_for_row(pseudo)

        new_point = {
            "Cycle_No": row["Cycle_No"],
            "Sample_No": row["Sample_No"],
            "Report_Date": row["Report_Date"],
            "Analyte": analyte_name,
            "Peer_Mean": row["Peer Mean"],
            "Your_Result": row["Your Result"],
            "%DEV": row["%DEV"],
            "SDI": row["SDI"],
            "Target_Score": row["Target Score"],
            "TEa_or_TDPA(%)": row["TDPA_limit_percent"],
            "Internal_TEa": row["Internal_TEa"],
            "diff_from_mean": row["diff_from_mean"],
            "Within_Internal_TEa?": row["Within_Internal_TEa?"],  # ✅ ADD THIS
        }

        new_point_df = pd.DataFrame([new_point])

        # If there's no prior history for this analyte, just use the new row
        if sub_hist.empty or sub_hist.dropna(how="all").empty:
            simulated_hist = new_point_df.copy()
        else:
            simulated_hist = pd.concat(
                [sub_hist, new_point_df],
                ignore_index=True,
                sort=False,
                copy=False,
            )

        simulated_hist = simulated_hist.sort_values(["Report_Date"]).reset_index(drop=True)

        # STEP 1: base risk for this single cycle
        base_risk_label = base_risk_for_row(row)

        # We'll set this base risk to last row (i.e. this new cycle),
        # then escalate using historical context in simulated_hist
        simulated_hist["Risk_Category_BaseOnly"] = None
        simulated_hist.loc[simulated_hist.index[-1], "Risk_Category_BaseOnly"] = base_risk_label

        final_risk_label, bias_last3, trend_last3 = escalate_for_history(
            simulated_hist,
            analyte_name,
            row,
            base_risk_label,
        )

        comment_txt, action_txt = build_comment_and_action(final_risk_label, bias_last3, trend_last3)

        last = simulated_hist.index[-1]
        simulated_hist.loc[last, "Risk_Category_Final"] = final_risk_label
        simulated_hist.loc[last, "Bias_Flag_Last3"] = bool(bias_last3)
        simulated_hist.loc[last, "Trend_Flag_Last3"] = bool(trend_last3)
        simulated_hist.loc[last, "Required_Action"] = action_txt
        simulated_hist.loc[last, "Comment"] = comment_txt


        # harvest just the *new row* (the last one we added)
        new_hist_row = simulated_hist.iloc[[-1]]
        enriched_rows_for_history.append(new_hist_row)

        # also collect info for populating Result Summary sheet
        summary_eval_rows.append({
            "Parameter_Name": analyte_name,
            "Risk_Category_Final": final_risk_label,
            "Comment": comment_txt,
            "Required_Action": action_txt,
            "TDPA_limit_percent": row["TDPA_limit_percent"],
        })

    if not enriched_rows_for_history:
        print("  No new rows to add (likely already in Cycle_History). Still refreshing Latest_Cycle.")

        # ✅ still backfill TEa columns for existing history rows
        backfill_internal_tea_in_cycle_history(wb, internal_tea_map)

        update_latest_cycle_sheet(wb)
        wb.save(out_path)
        return

    enriched_rows_for_history_df = pd.concat(enriched_rows_for_history, ignore_index=True)

    # --- write Header Information sheet ---
    fill_header_information_sheet(wb, meta)

    # --- write Result Summary sheet (for THIS cycle) ---
    summary_eval_df = pd.DataFrame(summary_eval_rows)
    update_result_summary_sheet(
        wb,
        current_cycle_rows=df,
        final_eval=summary_eval_df
    )

    # --- append to Cycle_History sheet (for ALL cycles, longitudinal audit trail) ---
    update_cycle_history_sheet(
        wb,
        meta,
        enriched_rows_for_history_df[
            [
                "Cycle_No",
                "Sample_No",
                "Report_Date",
                "Analyte",
                "Peer_Mean",
                "Your_Result",
                "%DEV",
                "SDI",
                "Target_Score",
                "TEa_or_TDPA(%)",
                "Internal_TEa",  # ✅
                "diff_from_mean",
                "Risk_Category_BaseOnly",
                "Risk_Category_Final",
                "Bias_Flag_Last3",
                "Trend_Flag_Last3",
                "Within_Internal_TEa?",  # ✅
                "Required_Action",
                "Comment",
            ]
        ]
    )

    # ✅ refresh the reviewer sheet
    update_latest_cycle_sheet(wb)

    # --- save workbook ---
    wb.save(out_path)

############################################
# 5. GUI LAUNCHER (file pickers for PDFs / template / output)
############################################
if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from pathlib import Path

    def run_gui():
        root = tk.Tk()
        root.withdraw()
        root.update_idletasks()

        # 1) Select one or more RIQAS PDFs
        pdf_paths = filedialog.askopenfilenames(
            title="Select RIQAS PDF file(s)",
            filetypes=[("PDF files", "*.pdf")],
        )
        if not pdf_paths:
            messagebox.showinfo("Cancelled", "No PDFs selected.")
            return

        # 2) Select the Excel template
        template_path = filedialog.askopenfilename(
            title="Select Excel Template (.xlsx)",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not template_path:
            messagebox.showinfo("Cancelled", "No Excel template selected.")
            return

        # ============================================================
        # ✅ NEW STEP: Select the Internal TEa table ONCE (for all PDFs)
        # Put this RIGHT HERE (after template, before output folder)
        # ============================================================
        tea_path = filedialog.askopenfilename(
            title="Select Internal TEa table (xlsx or csv)",
            filetypes=[("Excel or CSV", "*.xlsx *.xls *.csv")],
        )
        if not tea_path:
            messagebox.showinfo("Cancelled", "No Internal TEa file selected.")
            return

        tea_map = load_internal_tea_map(Path(tea_path))
        # ============================================================

        # 3) Choose output folder
        out_dir = filedialog.askdirectory(
            title="Select output folder for generated .xlsx files"
        )
        if not out_dir:
            messagebox.showinfo("Cancelled", "No output folder selected.")
            return

        # Normalize paths
        pdf_list = [Path(p) for p in root.tk.splitlist(pdf_paths)]
        template = Path(template_path)
        out_dir = Path(out_dir)

        # --- SORT PDFs BY REPORT DATE ---
        from datetime import date

        def get_pdf_report_date(pdf_path: Path):
            txt = read_pdf_text(pdf_path)
            meta = parse_metadata(txt)
            return meta.get("report_date") or date.min

        pdf_list_sorted = sorted(pdf_list, key=get_pdf_report_date)

        # ============================================================
        # ✅ GUI STEP 3 — CREATE ROLLING WORKBOOK ONCE
        # ============================================================
        import shutil

        rolling_out = out_dir / "RIQAS_EQA_Rolling_History.xlsx"

        if not rolling_out.exists():
            shutil.copyfile(template, rolling_out)
        # ============================================================

        # 4) Process each selected PDF (IN DATE ORDER)
        errors = []
        ok_count = 0

        for pdf in pdf_list_sorted:
            print("Processing:", pdf)  # ✅ correct place

            try:
                process_riqas_pdf_into_workbook(
                    pdf_path=str(pdf),
                    xlsx_path=str(rolling_out),
                    out_path=str(rolling_out),
                    internal_tea_map=tea_map,  # ✅ add this
                )

                ok_count += 1
            except Exception as e:
                errors.append(f"{pdf.name}: {e!r}")

        # 5) Result dialog
        if errors:
            msg_lines = [
                f"Created {ok_count} / {len(pdf_list)} workbook(s) in:",
                str(out_dir),
                "",
                "Errors:",
                *errors
            ]
            messagebox.showerror("Done with errors", "\n".join(msg_lines[:50]))
        else:
            messagebox.showinfo(
                "Success",
                f"Created {ok_count} workbook(s) in:\n{out_dir}"
            )

    run_gui()
